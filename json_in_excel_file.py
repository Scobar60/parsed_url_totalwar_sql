import json  # подключили библиотеку для работы с json
from pprint import pprint  # подключили Pprint для красоты выдачи текста
import openpyxl

with open ('result.json', 'r', encoding='utf-8') as file:
    text = json.load(file)



book = openpyxl.Workbook()

sheet = book.active

sheet.cell(row=1, column=1).value = 'Дата'
sheet.cell(row=1, column=2).value = 'Название статьи'
sheet.cell(row=1, column=3).value = "Текст статьи"
sheet.cell(row=1, column=4).value = "Ссылка на статью"
sheet.cell(row=1, column=5).value = "Ссылка на рисунок"


row = 2
count_row = len(text)-19001
print(count_row)
i = 0
for i in range(count_row):

    sheet[row][0].value = text[i]['article_data']
    sheet[row][1].value = text[i]['article_title']
    sheet[row][2].value = text[i]['article_text']
    sheet[row][3].value = text[i]['url']
    sheet[row][4].value = text[i]['article_img']

    print(f'Обработал {i}/ {count_row}')

    i+= 1
    row += 1



book.save('my_book.xlsx')
book.close()
print("Завершено")


